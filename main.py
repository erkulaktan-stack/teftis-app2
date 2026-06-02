import flet as ft
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import shutil

def main(page: ft.Page):
    page.title = "ÖZDEMİRTEKS Tablet Teftiş Paneli"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.scroll = ft.ScrollMode.AUTO
    
    page.window_width = 850
    page.window_height = 850
    
    # Başlangıçta beyaz ekranı önlemek için geçici bir yükleniyor ekranı basıyoruz
    durum_mesaji = ft.Text("Sistem başlatılıyor ve dosya yolları kontrol ediliyor...", size=16, weight=ft.FontWeight.BOLD)
    page.add(ft.Container(content=durum_mesaji, padding=20, alignment=ft.alignment.center))
    page.update()

    # --- SAYAÇLAR ---
    sayaclar = {"major": 0, "minor": 0}
    major_txt = ft.Text(value="0", size=28, weight=ft.FontWeight.BOLD, color=ft.Colors.RED_800)
    minor_txt = ft.Text(value="0", size=28, weight=ft.FontWeight.BOLD, color=ft.Colors.ORANGE_800)

    def sayac_degistir(islem, tip):
        if islem == "artir":
            sayaclar[tip] += 1
        elif islem == "azalt" and sayaclar[tip] > 0:
            sayaclar[tip] -= 1
        major_txt.value = str(sayaclar["major"])
        minor_txt.value = str(sayaclar["minor"])
        page.update()

    # --- RESİM SEÇİM SÖZLÜKLERİ ---
    secilen_resimler = {
        "olcu_xs": None, "olcu_s": None, "olcu_m": None, "olcu_l": None, "olcu_xl": None, "olcu_xxl": None,
        "minor_img": None, "major_img": None, "onden": None, "arkadan": None, "paketli_on": None, "paketli_arka": None,
        "sirali_acik": None, "tek_acik": None, "kapali_sirali": None, "xs_img": None, "s_img": None, "m_img": None,
        "l_img": None, "xl_img": None, "xxl_img": None
    }

    onizleme_yazilari = {k: ft.Text("Seçilmedi", size=12, italic=True, color=ft.Colors.GREY_600) for k in secilen_resimler.keys()}

    aktif_kategori = [None]

    def dosya_secildi_sonuc(e: ft.FilePickerResultEvent):
        if e.files and aktif_kategori[0]:
            dosya_yolu = e.files[0].path
            kat = aktif_kategori[0]
            secilen_resimler[kat] = dosya_yolu
            onizleme_yazilari[kat].value = os.path.basename(dosya_yolu)
            onizleme_yazilari[kat].color = ft.Colors.GREEN_700
            onizleme_yazilari[kat].weight = ft.FontWeight.BOLD
            page.update()

    file_picker = ft.FilePicker(on_result=dosya_secildi_sonuc)
    page.overlay.append(file_picker)

    def tetikle_dosya_sec(kategori):
        aktif_kategori[0] = kategori
        file_picker.pick_files(allow_multiple=False, file_type=ft.FilePickerFileType.IMAGE)

    def resim_secici_butonu(etiket, kategori):
        return ft.Row([
            ft.Button(
                content=ft.Row([
                    ft.Icon(ft.Icons.CAMERA_ALT, color=ft.Colors.BLACK),
                    ft.Text(value=etiket, color=ft.Colors.BLACK, weight=ft.FontWeight.W_500)
                ], alignment=ft.MainAxisAlignment.CENTER, spacing=5),
                on_click=lambda e: tetikle_dosya_sec(kategori),
                width=260,
                style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_GREY_100)
            ),
            onizleme_yazilari[kategori]
        ], alignment=ft.MainAxisAlignment.START, spacing=10)

    # --- INPUT ALANLARI VE FORM ELEMANLARI ---
    def hesapla_aql(e):
        try:
            qty = int(order_qty_input.value) if order_qty_input.value else 0
            sample_size, max_major, max_minor = 0, 0, 0
            if 151 <= qty <= 280: sample_size, max_major, max_minor = 32, 2, 3
            elif 281 <= qty <= 500: sample_size, max_major, max_minor = 50, 3, 5
            elif 501 <= qty <= 1200: sample_size, max_major, max_minor = 80, 5, 7
            elif 1201 <= qty <= 3200: sample_size, max_major, max_minor = 125, 7, 10
            elif 3201 <= qty <= 10000: sample_size, max_major, max_minor = 200, 10, 14
            
            sample_size_txt.value = f"Numune Miktarı: {sample_size} Adet"
            aql_limit_txt.value = f"Kabul Sınırı -> Maks Major: {max_major} | Maks Minor: {max_minor}"
        except ValueError:
            sample_size_txt.value = "Hata"
        page.update()

    brand_dropdown = ft.Dropdown(label="Marka (Brand)", options=[ft.dropdown.Option("JD SPORT"), ft.dropdown.Option("Gymking"), ft.dropdown.Option("Marshall")], value="JD SPORT", width=330)
    supplier_input = ft.TextField(label="Tedarikçi (Supplier)", value="OZDEMIRTEKS TEKSTIL", width=330)
    factory_input = ft.TextField(label="Fabrika (Factory)", value="OZDEMIRTEKS TEKSTIL", width=330)
    inspector_input = ft.TextField(label="Denetçi (Checked by)", value="Merdali Mete", width=330)
    po_input = ft.TextField(label="PO Numarası", value="none", width=330)
    style_name_input = ft.TextField(label="Stil Adı (Style Name)", value="JD SPORT", width=330)
    style_num_input = ft.TextField(label="Stil No (Style Number)", value="none", width=330)
    order_qty_input = ft.TextField(label="Sipariş Miktarı (Order Qty)", value="2688", width=330, on_change=hesapla_aql)
    shipped_qty_input = ft.TextField(label="Sevk Miktarı (Shipped Qty)", value="2688", width=330)
    colour_input = ft.TextField(label="Renk (Colour)", value="none", width=330)

    sample_size_txt = ft.Text(value="Numune Miktarı: 125 Adet", size=16, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_800)
    aql_limit_txt = ft.Text(value="Kabul Sınırı -> Maks Major: 7 | Maks Minor: 10", size=14, color=ft.Colors.BLUE_GREY_700)

    form_sol = ft.Column([brand_dropdown, supplier_input, factory_input, inspector_input, colour_input])
    form_sag = ft.Column([po_input, style_name_input, style_num_input, order_qty_input, shipped_qty_input])
    sonuc_kutusu = ft.Container(content=ft.Column([sample_size_txt, aql_limit_txt]), padding=15, bgcolor=ft.Colors.BLUE_GREY_50, border_radius=8, width=700)

    def ekrani_degistir(e):
        ekran_1.visible = False
        ekran_2.visible = True
        page.update()

    def ana_ekrana_don(e):
        ekran_2.visible = False
        ekran_1.visible = True
        page.update()

    buton_sonraki = ft.Button(
        content=ft.Text("Sonraki Adım: Resim ve Hata Girişi", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD), 
        bgcolor=ft.Colors.BLUE_800, 
        width=320, 
        height=50, 
        on_click=ekrani_degistir
    )

    baslik_alani = ft.Row([
        ft.Text("ÖZDEMİRTEKS - INSPECTION PROGRAMI", size=20, weight=ft.FontWeight.BOLD)
    ], alignment=ft.MainAxisAlignment.START, spacing=15)

    ekran_1 = ft.Column(controls=[
        baslik_alani,
        ft.Divider(), 
        ft.Row([form_sol, form_sag], spacing=20, alignment=ft.MainAxisAlignment.START), 
        ft.Divider(),
        ft.Text("MÜŞTERİ SAMPLING PLAN STANDARTLARI", size=16, weight=ft.FontWeight.BOLD),
        sonuc_kutusu, ft.Divider(), ft.Row([buton_sonraki], alignment=ft.MainAxisAlignment.CENTER)
    ], visible=True)

    # --- EKRAN 2: RESİMLER VE SAYAÇLAR ---
    bilgi_notu = ft.Text(value="", size=14, weight=ft.FontWeight.BOLD, color=ft.Colors.GREEN_700)

    def orjinal_excele_kaydet(e):
        # Verdiğiniz harika yol bilgisine göre Android çift yol kontrolü uyguluyoruz
        sablon_yolu1 = "/sdcard/Download/INSPECTION REPORT.xlsx"
        sablon_yolu2 = "/storage/emulated/0/Download/INSPECTION REPORT.xlsx"
        
        sablon_yolu = sablon_yolu1 if os.path.exists(sablon_yolu1) else sablon_yolu2
        cikti_yolu = f"/sdcard/Download/OZDEMIRTEKS_TEFTIS_RAPORU_PO_{po_input.value}.xlsx"
        
        if not os.path.exists(sablon_yolu):
            bilgi_notu.value = f"Hata: Tabletin Dahili Depolama > Download klasöründe 'INSPECTION REPORT.xlsx' bulunamadı!"
            bilgi_notu.color = ft.Colors.RED_700
            page.update()
            return

        try:
            bilgi_notu.value = "Rapor hazırlanıyor, fotoğraflar işleniyor..."
            bilgi_notu.color = ft.Colors.BLUE_700
            page.update()

            shutil.copy(sablon_yolu, cikti_yolu)
            wb = load_workbook(cikti_yolu)
            
            def sekme_bul(isim):
                for s in wb.sheetnames:
                    if s.strip().upper() == isim.strip().upper(): return s
                for s in wb.sheetnames:
                    if isim.strip().upper() in s.strip().upper(): return s
                return None

            paperwork_sekme = sekme_bul("AQL - PAPERWORK COPY")
            if paperwork_sekme:
                ws1 = wb[paperwork_sekme]
                ws1["A10"] = f"PO: {po_input.value}"
                ws1["A11"] = f"STYLE NAME:  {style_name_input.value}"
                ws1["A12"] = f"STYLE NUMBER:{style_num_input.value}"
                ws1["A13"] = f"SUPPLIER: {supplier_input.value}"
                ws1["A14"] = f"FACTORY: {factory_input.value}"
                ws1["A15"] = f"Checked by: {inspector_input.value}"
                ws1["D12"] = int(order_qty_input.value) if order_qty_input.value.isdigit() else order_qty_input.value
                ws1["E9"] = sayaclar["major"]
                ws1["F9"] = sayaclar["minor"]
                ws1["A16"] = f"COLOUR: {colour_input.value}"
                ws1["D13"] = int(shipped_qty_input.value) if shipped_qty_input.value.isdigit() else shipped_qty_input.value  

            resim_haritasi = [
                ("olcu_xs", "AQL MEASUREMENT CHART COPY", "A7"),
                ("olcu_s", "AQL MEASUREMENT CHART COPY", "P7"),
                ("olcu_m", "AQL MEASUREMENT CHART COPY", "A27"),
                ("olcu_l", "AQL MEASUREMENT CHART COPY", "P27"),
                ("olcu_xl", "AQL MEASUREMENT CHART COPY", "A47"),
                ("olcu_xxl", "AQL MEASUREMENT CHART COPY", "P47"),
                
                ("minor_img", "Photos - AQL COPY", "C3"),
                ("major_img", "Photos - AQL COPY", "I3"),
                ("onden", "Photos - AQL COPY", "B20"),
                ("arkadan", "Photos - AQL COPY", "C20"),
                ("paketli_on", "Photos - AQL COPY", "I20"),
                ("paketli_arka", "Photos - AQL COPY", "O20"), # typo düzeltildi
                ("sirali_acik", "Photos - AQL COPY", "B38"),
                ("tek_acik", "Photos - AQL COPY", "C38"),
                ("kapali_sirali", "Photos - AQL COPY", "I38"),
                ("xs_img", "Photos - AQL COPY", "B56"),
                ("s_img", "Photos - AQL COPY", "C56"),
                ("m_img", "Photos - AQL COPY", "I56"),
                ("l_img", "Photos - AQL COPY", "O56"),
                ("xl_img", "Photos - AQL COPY", "B70"),
                ("xxl_img", "Photos - AQL COPY", "C70"),
            ]

            for anahtar, hedef_kelime, hucre in resim_haritasi:
                resim_yolu = secilen_resimler[anahtar]
                if resim_yolu and os.path.exists(resim_yolu):
                    gercek_sekme_adi = sekme_bul(hedef_kelime)
                    if gercek_sekme_adi:
                        ws_resim = wb[gercek_sekme_adi]
                        img = OpenpyxlImage(resim_yolu)
                        
                        # Hücre boyutlarına tam oturması için otomatik boyutlandırma
                        if "MEASUREMENT" in gercek_sekme_adi.upper():
                            img.width, img.height = 240, 320
                        else:
                            img.width, img.height = 280, 210
                            
                        ws_resim.add_image(img, hucre)

            wb.save(cikti_yolu)
            bilgi_notu.value = f"Başarılı! Rapor 'Download' klasörüne kaydedildi.\nDosya: OZDEMIRTEKS_TEFTIS_RAPORU_PO_{po_input.value}.xlsx"
            bilgi_notu.color = ft.Colors.GREEN_700
        except Exception as ex:
            bilgi_notu.value = f"Rapor kaydedilirken hata oluştu: {str(ex)}"
            bilgi_notu.color = ft.Colors.RED_700
        page.update()

    # Ekran 2 Elemanları (Sayaç ve Resim Paneli)
    sayac_paneli = ft.Row([
        ft.Container(content=ft.Column([
            ft.Text("MAJOR HATA", size=14, weight=ft.FontWeight.BOLD),
            ft.Row([
                ft.IconButton(ft.Icons.REMOVE_CIRCLE, on_click=lambda e: sayac_degistir("azalt", "major"), icon_color=ft.Colors.RED_700),
                major_txt,
                ft.IconButton(ft.Icons.ADD_CIRCLE, on_click=lambda e: sayac_degistir("artir", "major"), icon_color=ft.Colors.RED_700)
            ], alignment=ft.MainAxisAlignment.CENTER)
        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER), padding=10, border=ft.border.all(1, ft.Colors.RED_200), border_radius=8, width=200),
        
        ft.Container(content=ft.Column([
            ft.Text("MINOR HATA", size=14, weight=ft.FontWeight.BOLD),
            ft.Row([
                ft.IconButton(ft.Icons.REMOVE_CIRCLE, on_click=lambda e: sayac_degistir("azalt", "minor"), icon_color=ft.Colors.ORANGE_700),
                minor_txt,
                ft.IconButton(ft.Icons.ADD_CIRCLE, on_click=lambda e: sayac_degistir("artir", "minor"), icon_color=ft.Colors.ORANGE_700)
            ], alignment=ft.MainAxisAlignment.CENTER)
        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER), padding=10, border=ft.border.all(1, ft.Colors.ORANGE_200), border_radius=8, width=200)
    ], spacing=30, alignment=ft.MainAxisAlignment.CENTER)

    resim_listesi = ft.Column([
        ft.Text("Ölçüm Tablosu Fotoğrafları (AQL MEASUREMENT)", weight=ft.FontWeight.BOLD, size=14),
        resim_secici_butonu("XS Ölçüm Kartı", "olcu_xs"),
        resim_secici_butonu("S Ölçüm Kartı", "olcu_s"),
        resim_secici_butonu("M Ölçüm Kartı", "olcu_m"),
        resim_secici_butonu("L Ölçüm Kartı", "olcu_l"),
        resim_secici_butonu("XL Ölçüm Kartı", "olcu_xl"),
        resim_secici_butonu("XXL Ölçüm Kartı", "olcu_xxl"),
        ft.Divider(),
        ft.Text("Genel Teftiş Fotoğrafları (Photos - AQL)", weight=ft.FontWeight.BOLD, size=14),
        resim_secici_butonu("Minor Hata Görseli", "minor_img"),
        resim_secici_butonu("Major Hata Görseli", "major_img"),
        resim_secici_butonu("Ürün Önden Görünüm", "onden"),
        resim_secici_butonu("Ürün Arkadan Görünüm", "arkadan"),
        resim_secici_butonu("Paketli Ön Görünüm", "paketli_on"),
        resim_secici_butonu("Paketli Arka Görünüm", "paketli_arka"),
        resim_secici_butonu("Sıralı Açık Görünüm", "sirali_acik"),
        resim_secici_butonu("Tek Açık Görünüm", "tek_acik"),
        resim_secici_butonu("Kapalı Sıralı Görünüm", "kapali_sirali"),
        resim_secici_butonu("XS Ürün Fotoğrafı", "xs_img"),
        resim_secici_butonu("S Ürün Fotoğrafı", "s_img"),
        resim_secici_butonu("M Ürün Fotoğrafı", "m_img"),
        resim_secici_butonu("L Ürün Fotoğrafı", "l_img"),
        resim_secici_butonu("XL Ürün Fotoğrafı", "xl_img"),
        resim_secici_butonu("XXL Ürün Fotoğrafı", "xxl_img"),
    ], spacing=10)

    ekran_2 = ft.Column(controls=[
        ft.Row([
            ft.IconButton(ft.Icons.ARROW_BACK, on_click=ana_ekrana_don),
            ft.Text("Hata Sayacı ve Resim Ekleme Paneli", size=18, weight=ft.FontWeight.BOLD)
        ]),
        ft.Divider(),
        sayac_paneli,
        ft.Divider(),
        resim_listesi,
        ft.Divider(),
        ft.Row([
            ft.ElevatedButton(
                content=ft.Text("EXCEL RAPORUNU OLUŞTUR", color=ft.Colors.WHITE, weight=ft.FontWeight.BOLD),
                bgcolor=ft.Colors.GREEN_700,
                width=350,
                height=55,
                on_click=orjinal_excele_kaydet
            )
        ], alignment=ft.MainAxisAlignment.CENTER),
        ft.Row([bilgi_notu], alignment=ft.MainAxisAlignment.CENTER)
    ], visible=False)

    # Başlangıç yükleniyor yazısını silip ana ekranları ekliyoruz
    page.clean()
    page.add(ekran_1, ekran_2)
    page.update()

ft.app(target=main)
